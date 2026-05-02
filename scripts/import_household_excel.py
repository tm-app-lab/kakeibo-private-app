from __future__ import annotations

import json
import csv
import re
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries


DEFAULT_EXCEL_PATH = Path("C:/Users/manab/Desktop") / "\u5bb6\u8a08.xlsx"
DESKTOP_PATH = Path("C:/Users/manab/Desktop")
OUTPUT_PATH = Path(__file__).resolve().parents[1] / "data" / "household.json"
JS_OUTPUT_PATH = Path(__file__).resolve().parents[1] / "data" / "household-data.js"
MONEY_FORWARD_FILES = [
    DESKTOP_PATH / "\u53ce\u5165\u30fb\u652f\u51fa\u8a73\u7d30_2025-04-01_2025-04-30.csv",
    DESKTOP_PATH / "\u53ce\u5165\u30fb\u652f\u51fa\u8a73\u7d30_2025-05-01_2025-05-31.csv",
    DESKTOP_PATH / "\u53ce\u5165\u30fb\u652f\u51fa\u8a73\u7d30_2025-06-01_2025-06-30.csv",
    DESKTOP_PATH / "\u53ce\u5165\u30fb\u652f\u51fa\u8a73\u7d30_2025-07-01_2025-07-31.csv",
    DESKTOP_PATH / "\u53ce\u5165\u30fb\u652f\u51fa\u8a73\u7d30_2025-08-01_2025-08-31.csv",
    DESKTOP_PATH / "\u53ce\u5165\u30fb\u652f\u51fa\u8a73\u7d30_2025-09-01_2025-09-30.csv",
    DESKTOP_PATH / "\u53ce\u5165\u30fb\u652f\u51fa\u8a73\u7d30_2025-10-01_2025-10-31.csv",
    DESKTOP_PATH / "\u53ce\u5165\u30fb\u652f\u51fa\u8a73\u7d30_2025-11-01_2025-11-30.csv",
    DESKTOP_PATH / "\u53ce\u5165\u30fb\u652f\u51fa\u8a73\u7d30_2025-12-01_2025-12-31.csv",
    DESKTOP_PATH / "\u53ce\u5165\u30fb\u652f\u51fa\u8a73\u7d30_2026-01-01_2026-01-31.csv",
    DESKTOP_PATH / "\u53ce\u5165\u30fb\u652f\u51fa\u8a73\u7d30_2026-02-01_2026-02-28.csv",
    DESKTOP_PATH / "\u53ce\u5165\u30fb\u652f\u51fa\u8a73\u7d30_2026-03-01_2026-03-31.csv",
    DESKTOP_PATH / "\u53ce\u5165\u30fb\u652f\u51fa\u8a73\u7d30_2026-04-01_2026-04-30.csv",
]
RAKUTEN_CARD_FILES = [
    DESKTOP_PATH / f"enavi{yyyymm}(4227).csv"
    for yyyymm in [
        "202604",
        "202603",
        "202602",
        "202601",
        "202512",
        "202511",
        "202510",
        "202509",
        "202508",
        "202507",
        "202506",
        "202505",
        "202504",
    ]
]

ALIAS_RULES = {
    "東京電力": ["東京電力", "トウキヨウデンリヨク", "トウキョウデンリョク"],
    "東電生協": ["東電生協", "トウデンセイキヨウ", "トウデンセイキョウ"],
    "KDDI": ["KDDI", "ケイディーディーアイ"],
    "JAF": ["JAF"],
    "Amazon": ["AMAZON", "Amazon", "アマゾン"],
    "住宅ローン": ["住宅ローン", "ローン"],
    "NHK": ["NHK"],
    "ENEOS": ["ENEOS", "エネオス"],
}


def clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, str):
        return value.strip()
    return value


def table_rows(ws, table_name: str) -> tuple[list[str], list[dict[str, Any]]]:
    table = ws.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    headers = [clean(ws.cell(min_row, col).value) or f"列{col}" for col in range(min_col, max_col + 1)]
    rows: list[dict[str, Any]] = []
    for row_index in range(min_row + 1, max_row + 1):
        record = {
            str(headers[col_index - min_col]): clean(ws.cell(row_index, col_index).value)
            for col_index in range(min_col, max_col + 1)
        }
        if any(value not in (None, "") for value in record.values()):
            rows.append(record)
    return [str(header) for header in headers], rows


def yen(value: Any) -> int:
    if isinstance(value, (int, float)):
        return int(round(value))
    return 0


def read_money_forward_csv(path: Path) -> list[dict[str, str]]:
    text = path.read_text(encoding="cp932")
    return list(csv.DictReader(text.splitlines()))


def read_rakuten_csv(path: Path) -> list[dict[str, str]]:
    text = path.read_text(encoding="utf-8-sig")
    return list(csv.DictReader(text.splitlines()))


def parse_money(value: Any) -> int:
    if value is None:
        return 0
    return int(str(value).replace(",", "").strip() or 0)


def month_key(date_text: str) -> str:
    return datetime.strptime(date_text, "%Y/%m/%d").strftime("%Y-%m")


def normalize_text(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).upper()
    return re.sub(r"\s+", "", text)


def item_id(item: dict[str, Any]) -> str:
    return "::".join(
        normalize_text(item.get(key))
        for key in ["person", "category", "name", "detail", "costType", "flow"]
    )


def token_set(*values: Any) -> set[str]:
    text = " ".join(str(value or "") for value in values)
    tokens = {normalize_text(token) for token in re.split(r"[\s　/・（）()【】\[\]、,.:：\-]+", text)}
    return {token for token in tokens if len(token) >= 2}


def alias_hit(excel_item: dict[str, Any], evidence: dict[str, Any]) -> bool:
    item_text = normalize_text(excel_item.get("name"))
    evidence_text = normalize_text(
        " ".join(str(evidence.get(key) or "") for key in ["content", "category", "middle", "institution"])
    )
    for label, aliases in ALIAS_RULES.items():
        normalized_aliases = [normalize_text(alias) for alias in aliases]
        if normalize_text(label) in item_text and any(alias in evidence_text for alias in normalized_aliases):
            return True
    return False


def amount_score(planned: int, actual: int) -> int:
    if planned <= 0 or actual <= 0:
        return 0
    ratio = abs(planned - actual) / max(planned, actual)
    if ratio <= 0.03:
        return 35
    if ratio <= 0.08:
        return 28
    if ratio <= 0.15:
        return 20
    if ratio <= 0.3:
        return 10
    return 0


def score_alignment(excel_item: dict[str, Any], evidence: dict[str, Any]) -> tuple[int, list[str]]:
    score = 0
    reasons: list[str] = []
    if excel_item["category"] == evidence["category"]:
        score += 25
        reasons.append("分類一致")
    if excel_item["costType"] == "固定" and evidence["months"] >= 4:
        score += 15
        reasons.append("毎月性あり")
    if alias_hit(excel_item, evidence):
        score += 35
        reasons.append("件名・別名一致")

    excel_tokens = token_set(excel_item.get("name"), excel_item.get("detail"))
    evidence_tokens = token_set(evidence.get("content"), evidence.get("middle"), evidence.get("institution"))
    overlap = excel_tokens & evidence_tokens
    if overlap:
        score += min(25, 10 + 5 * len(overlap))
        reasons.append("語句一致")

    value_score = max(amount_score(excel_item["amount"], evidence["average"]), amount_score(excel_item["amount"], evidence["latest"]))
    if value_score:
        score += value_score
        reasons.append("金額近似")

    return min(score, 100), reasons


def build_money_forward_actuals() -> dict[str, Any]:
    transactions: list[dict[str, Any]] = []
    imported_files: list[str] = []
    for path in MONEY_FORWARD_FILES:
        if not path.exists():
            continue
        imported_files.append(path.name)
        for row in read_money_forward_csv(path):
            is_counted = row.get("計算対象") == "1"
            is_transfer = row.get("振替") == "1"
            amount = parse_money(row.get("金額（円）"))
            signed_type = "income" if amount > 0 else "expense"
            major = row.get("大項目") or "未分類"
            middle = row.get("中項目") or "未分類"
            content = row.get("内容") or ""
            is_card_settlement = major == "現金・カード" and ("カード引き落とし" in middle or content == "クレジット")
            transactions.append(
                {
                    "date": row.get("日付"),
                    "month": month_key(row.get("日付")),
                    "content": content,
                    "amount": amount,
                    "absAmount": abs(amount),
                    "institution": row.get("保有金融機関") or "",
                    "major": major,
                    "middle": middle,
                    "memo": row.get("メモ") or "",
                    "isCounted": is_counted,
                    "isTransfer": is_transfer,
                    "isInternalMovement": is_transfer or is_card_settlement,
                    "type": signed_type,
                    "id": row.get("ID") or "",
                }
            )

    effective = [tx for tx in transactions if tx["isCounted"] and not tx["isInternalMovement"]]
    months = sorted({tx["month"] for tx in effective})
    monthly: list[dict[str, Any]] = []
    category_by_month: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    middle_by_month: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    institution_totals: dict[str, int] = defaultdict(int)

    for month in months:
        month_rows = [tx for tx in effective if tx["month"] == month]
        income = sum(tx["amount"] for tx in month_rows if tx["amount"] > 0)
        expense = sum(tx["absAmount"] for tx in month_rows if tx["amount"] < 0)
        for tx in month_rows:
            if tx["amount"] < 0:
                category_by_month[month][tx["major"]] += tx["absAmount"]
                middle_by_month[month][f"{tx['major']} / {tx['middle']}"] += tx["absAmount"]
                institution_totals[tx["institution"]] += tx["absAmount"]
        monthly.append(
            {
                "month": month,
                "income": income,
                "expense": expense,
                "balance": income - expense,
                "count": len(month_rows),
            }
        )

    latest_month = months[-1] if months else None
    latest_category = dict(category_by_month.get(latest_month, {}))
    average_category: dict[str, int] = {}
    for category in sorted({cat for values in category_by_month.values() for cat in values}):
        values = [category_by_month[month].get(category, 0) for month in months]
        average_category[category] = int(round(sum(values) / len(values))) if values else 0

    anomalies = []
    for category, latest_value in latest_category.items():
        average = average_category.get(category, 0)
        diff = latest_value - average
        if average and latest_value >= average * 1.25 and diff >= 10000:
            anomalies.append(
                {
                    "category": category,
                    "latest": latest_value,
                    "average": average,
                    "diff": diff,
                    "ratio": latest_value / average,
                }
            )
    anomalies.sort(key=lambda item: item["diff"], reverse=True)

    recurring_map: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for tx in effective:
        if tx["amount"] < 0:
            key = f"{tx['content']}|{tx['major']}|{tx['middle']}|{tx['institution']}"
            recurring_map[key].append(tx)
    expense_groups = []
    recurring_candidates = []
    for grouped in recurring_map.values():
        unique_months = sorted({tx["month"] for tx in grouped})
        amounts = [tx["absAmount"] for tx in grouped]
        latest_tx = sorted(grouped, key=lambda tx: tx["date"])[-1]
        group_record = {
            "content": latest_tx["content"],
            "category": latest_tx["major"],
            "middle": latest_tx["middle"],
            "institution": latest_tx["institution"],
            "months": len(unique_months),
            "average": int(round(sum(amounts) / len(amounts))),
            "latest": latest_tx["absAmount"],
            "min": min(amounts),
            "max": max(amounts),
            "total": sum(amounts),
            "lastDate": latest_tx["date"],
            "ids": [tx["id"] for tx in grouped],
        }
        expense_groups.append(group_record)
        if len(unique_months) >= 4:
            avg = int(round(sum(amounts) / len(amounts)))
            if avg >= 1000:
                recurring_candidates.append(
                    {
                        "content": group_record["content"],
                        "category": group_record["category"],
                        "middle": group_record["middle"],
                        "institution": group_record["institution"],
                        "months": len(unique_months),
                        "average": avg,
                        "min": min(amounts),
                        "max": max(amounts),
                    }
                )
    recurring_candidates.sort(key=lambda item: (item["months"], item["average"]), reverse=True)

    large_transactions = sorted(
        [tx for tx in effective if tx["amount"] < 0],
        key=lambda tx: tx["absAmount"],
        reverse=True,
    )[:30]

    return {
        "importedFiles": imported_files,
        "months": months,
        "monthly": monthly,
        "categoryByMonth": {month: dict(values) for month, values in category_by_month.items()},
        "middleByMonth": {month: dict(values) for month, values in middle_by_month.items()},
        "latestMonth": latest_month,
        "latestCategory": dict(sorted(latest_category.items(), key=lambda item: item[1], reverse=True)),
        "averageCategory": dict(sorted(average_category.items(), key=lambda item: item[1], reverse=True)),
        "institutionTotals": dict(sorted(institution_totals.items(), key=lambda item: item[1], reverse=True)),
        "anomalies": anomalies[:8],
        "expenseGroups": sorted(expense_groups, key=lambda item: (item["months"], item["total"]), reverse=True),
        "recurringCandidates": recurring_candidates[:20],
        "largeTransactions": large_transactions,
        "transactions": sorted(effective, key=lambda tx: tx["date"], reverse=True)[:400],
        "summary": {
            "months": len(months),
            "totalIncome": sum(row["income"] for row in monthly),
            "totalExpense": sum(row["expense"] for row in monthly),
            "averageIncome": int(round(sum(row["income"] for row in monthly) / len(monthly))) if monthly else 0,
            "averageExpense": int(round(sum(row["expense"] for row in monthly) / len(monthly))) if monthly else 0,
            "averageBalance": int(round(sum(row["balance"] for row in monthly) / len(monthly))) if monthly else 0,
            "transactionCount": len(effective),
            "excludedInternalMovements": len([tx for tx in transactions if tx["isCounted"] and tx["isInternalMovement"]]),
        },
    }


def build_rakuten_card_actuals() -> dict[str, Any]:
    rows: list[dict[str, Any]] = []
    for path in RAKUTEN_CARD_FILES:
        if not path.exists():
            continue
        payment_yyyymm = path.name[5:11]
        payment_month = f"{payment_yyyymm[:4]}-{payment_yyyymm[4:]}"
        for row in read_rakuten_csv(path):
            payment_key = next((key for key in row if key.endswith("月支払金額")), "支払総額")
            date = row.get("利用日") or ""
            rows.append(
                {
                    "sourceType": "rakuten",
                    "sourceFile": path.name,
                    "month": payment_month,
                    "useMonth": date.replace("/", "-")[:7],
                    "date": date,
                    "content": row.get("利用店名・商品名") or "",
                    "user": row.get("利用者") or "",
                    "paymentMethod": row.get("支払方法") or "",
                    "amount": row.get("利用金額") or "",
                    "fee": row.get("手数料/利息") or "",
                    "total": row.get("支払総額") or "",
                    "paymentAmount": row.get(payment_key) or row.get("支払総額") or row.get("利用金額") or "",
                    "carryover": next((row.get(key) for key in row if key.endswith("月繰越残高")), "") or "",
                    "sign": row.get("新規サイン") or "",
                    "id": f"{path.name}-{date}-{row.get('利用店名・商品名')}-{row.get('支払総額')}",
                }
            )
    months = sorted({row["month"] for row in rows}, reverse=True)
    return {"rows": rows, "months": months, "summary": {"rows": len(rows), "months": len(months)}}


def find_amount_key(headers: list[str]) -> str:
    amount_headers = [header for header in headers if "費用" in header]
    return amount_headers[-1] if amount_headers else headers[-1]


def build_master_alignment(excel_items: list[dict[str, Any]], money_forward: dict[str, Any]) -> dict[str, Any]:
    expense_items = [item for item in excel_items if item["flow"] == "支出"]
    evidence_groups = money_forward.get("expenseGroups", [])
    matched_evidence_keys: set[str] = set()
    update_candidates = []
    review_candidates = []
    unmatched_excel = []

    for item in expense_items:
        scored = []
        for evidence in evidence_groups:
            score, reasons = score_alignment(item, evidence)
            if score:
                scored.append((score, reasons, evidence))
        scored.sort(key=lambda row: row[0], reverse=True)
        if not scored:
            unmatched_excel.append({"itemId": item_id(item), "item": item, "reason": "候補なし"})
            continue

        best_score, reasons, evidence = scored[0]
        evidence_key = "|".join([evidence["content"], evidence["category"], evidence["middle"], evidence["institution"]])
        matched_evidence_keys.add(evidence_key)
        candidate = {
            "itemId": item_id(item),
            "excel": item,
            "moneyForward": evidence,
            "score": best_score,
            "confidence": "high" if best_score >= 75 else "medium" if best_score >= 50 else "low",
            "reasons": reasons,
            "suggestedAmount": evidence["average"] if item["costType"] == "固定" else evidence["latest"],
            "difference": (evidence["average"] if item["costType"] == "固定" else evidence["latest"]) - item["amount"],
        }
        has_direct_identity = "分類一致" in reasons or "語句一致" in reasons or "件名・別名一致" in reasons
        if best_score >= 75 and has_direct_identity and abs(candidate["difference"]) >= 100:
            update_candidates.append(candidate)
        elif best_score >= 45:
            review_candidates.append(candidate)
        else:
            unmatched_excel.append({"itemId": item_id(item), "item": item, "reason": "確度不足", "best": candidate})

    unmatched_money_forward = []
    for evidence in evidence_groups:
        evidence_key = "|".join([evidence["content"], evidence["category"], evidence["middle"], evidence["institution"]])
        if evidence_key in matched_evidence_keys:
            continue
        if evidence["total"] >= 10000 or evidence["months"] >= 3:
            unmatched_money_forward.append(evidence)

    return {
        "summary": {
            "excelExpenseItems": len(expense_items),
            "highConfidenceUpdates": len(update_candidates),
            "reviewCandidates": len(review_candidates),
            "unmatchedExcelItems": len(unmatched_excel),
            "unmatchedMoneyForwardItems": len(unmatched_money_forward),
        },
        "updateCandidates": sorted(update_candidates, key=lambda item: abs(item["difference"]), reverse=True),
        "reviewCandidates": sorted(review_candidates, key=lambda item: item["score"], reverse=True),
        "unmatchedExcel": unmatched_excel[:30],
        "unmatchedMoneyForward": sorted(unmatched_money_forward, key=lambda item: (item["months"], item["total"]), reverse=True)[:50],
    }


def build_dataset(excel_path: Path) -> dict[str, Any]:
    wb = load_workbook(excel_path, read_only=False, data_only=False)
    current_sheet_name = "R7.4" if "R7.4" in wb.sheetnames else wb.sheetnames[0]
    current_ws = wb[current_sheet_name]
    current_table_name = next(iter(current_ws.tables.keys()))
    headers, rows = table_rows(current_ws, current_table_name)
    amount_key = find_amount_key(headers)

    normalized: list[dict[str, Any]] = []
    for row in rows:
        amount = yen(row.get(amount_key))
        normalized.append(
            {
                "person": row.get("名前") or "未設定",
                "category": row.get("種別") or "未分類",
                "payment": row.get("支払い") or "未設定",
                "name": row.get("内容1") or "未設定",
                "detail": row.get("内容2") or "",
                "costType": row.get("固定/変動") or "未設定",
                "flow": row.get("フロー") or "支出",
                "amount": amount,
            }
        )

    totals = {
        "expense": sum(item["amount"] for item in normalized if item["flow"] == "支出"),
        "saving": sum(item["amount"] for item in normalized if item["flow"] == "貯蓄"),
        "fixed": sum(item["amount"] for item in normalized if item["costType"] == "固定" and item["flow"] == "支出"),
        "variable": sum(item["amount"] for item in normalized if item["costType"] == "変動" and item["flow"] == "支出"),
    }

    by_category: dict[str, int] = defaultdict(int)
    by_person: dict[str, int] = defaultdict(int)
    by_payment: dict[str, int] = defaultdict(int)
    for item in normalized:
        if item["flow"] == "支出":
            by_category[item["category"]] += item["amount"]
            by_person[item["person"]] += item["amount"]
            by_payment[item["payment"]] += item["amount"]

    incomes = [
        {"person": "孝", "label": "A口座への入金（給料）", "amount": yen(current_ws["J2"].value), "note": clean(current_ws["H2"].value)},
        {"person": "絵里香", "label": "A口座への入金（給料）", "amount": yen(current_ws["J6"].value), "note": clean(current_ws["H6"].value)},
    ]
    income_total = sum(item["amount"] for item in incomes)
    cash_after_outflow = income_total - totals["expense"] - totals["saving"]
    loan_amount = next((item["amount"] for item in normalized if item["name"] == "住宅ローン"), 0)
    money_forward = build_money_forward_actuals()
    master_alignment = build_master_alignment(normalized, money_forward)
    rakuten_card = build_rakuten_card_actuals()

    return {
        "source": str(excel_path),
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "period": current_sheet_name,
        "amountKey": amount_key,
        "incomes": incomes,
        "totals": {
            **totals,
            "income": income_total,
            "cashAfterOutflow": cash_after_outflow,
            "housingLoanRatio": loan_amount / income_total if income_total else 0,
        },
        "categoryTotals": dict(sorted(by_category.items(), key=lambda item: item[1], reverse=True)),
        "personTotals": dict(sorted(by_person.items(), key=lambda item: item[1], reverse=True)),
        "paymentTotals": dict(sorted(by_payment.items(), key=lambda item: item[1], reverse=True)),
        "items": sorted(normalized, key=lambda item: (item["flow"], item["category"], -item["amount"])),
        "importRules": [
            {"contains": "Amazon", "category": "日用品", "confidence": "medium"},
            {"contains": "ENEOS", "category": "車", "confidence": "high"},
            {"contains": "東京電力", "category": "光熱費", "confidence": "high"},
            {"contains": "NTT", "category": "通信", "confidence": "high"},
        ],
        "moneyForward": money_forward,
        "rakutenCard": rakuten_card,
        "masterAlignment": master_alignment,
    }


def main() -> None:
    dataset = build_dataset(DEFAULT_EXCEL_PATH)
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(dataset, ensure_ascii=False, indent=2), encoding="utf-8")
    JS_OUTPUT_PATH.write_text(
        "window.HOUSEHOLD_DATA = "
        + json.dumps(dataset, ensure_ascii=False, indent=2)
        + ";\n",
        encoding="utf-8",
    )
    print(f"Wrote {OUTPUT_PATH}")
    print(f"Wrote {JS_OUTPUT_PATH}")


if __name__ == "__main__":
    main()
